# -*- coding: utf-8 -*-
"""
Tools for a Farm's cowslist operation
    v1.0
    2022/7/30
    by jicc

"""
import openpyxl
import datetime
import chghistory
import fmstls

#fpyDF_cowslist#############################################################
"""
fpyDF_cowslist:
	Excelfile に 
    Umotionの　飼養牛一覧yyyymmdd から
    cowslistyyyymmdd を作成する
    v1.01
    2022/8/2
    @author:jicc

"""
#! python3
def fpyDF_cowslist( wbN, sheetorg, sheetN, fillinDate ):
    """
    Excelfile に 
    Umotionの　飼養牛一覧yyyymmdd から
    cowslistyyyymmdd を作成する
    
    Parameters
    ----------
    wbN : ワークブック名
        "AB_cowslist.xlsx"
    sheetorg : データ参照シート
        "cowslistyyyymmddorg"
    sheetN : 経産牛データシート
        "cowslistyyyymmdd"
    fillinDate : 作成日
        "yyyy/mm/dd"
    
    Returns
    -------
    None.

    """
    
    #import openpyxl
    #import datetime

    wb = openpyxl.load_workbook(wbN)
    sheetorg = wb[sheetorg]

    sheetN = wb[sheetN]
    max_row = sheetorg.max_row
    fillinDate = datetime.datetime.strptime(fillinDate, '%Y/%m/%d')
    fillinDate = fillinDate.date()  
    #date only yyyy-mm-dd v1.01 or + ".strftime('%Y/%m/%d')"
    
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
    
      #LineNo      
      sheetN.cell(row=row_num, column=1).value = row_num - 1
      
      #cowidNo
      sheetN.cell(row=row_num, column=2).value = sheetorg.cell(row=row_num, 
                                                            column=2).value
                                                  #個体識別番号
      #DHITNo
      sheetN.cell(row=row_num, column=4).value = sheetorg.cell(row=row_num, 
                                                            column=1).value
                                                  #牛番号
      #birthday
      birthday = sheetorg.cell(row=row_num, column=6).value
      birthday = birthday.date() 
      #date only yyyy-mm-dd v1.01 or + ".strftime('%Y/%m/%d')"
      sheetN.cell(row=row_num, column=7).value = birthday 
          #sheetorg.cell(row=row_num, column=6).value
                                                  #出生日
      #sire_code     
      sheetN.cell(row=row_num, column=9).value = sheetorg.cell(row=row_num, 
                                                            column=14).value
                                                  #父牛の略号
      #sire_name
      sheetN.cell(row=row_num, column=10).value = sheetorg.cell(row=row_num, 
                                                            column=13).value
                                                  #父牛の名前
      #fillindate
      sheetN.cell(row=row_num, column=20).value = fillinDate
      
                                                
    
    wb.save(wbN)

#fpyind_inf_to_cowlists#####################################################    
"""
fpyind_inf_to_cowslist:
    input individual informations from cowshistory's data
    to cowslist
    v1.0
    2022/8/2
    @author: jicc
    
    
"""

def fpyind_inf_to_cowslist(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1):
    """
    input individual informations from cowshistory's data
    to cowslist    

    Parameters
    ----------
    wbN0 : str
        Excelfile  data of trans information
        ex. "AB_cowhistory.xlsx"
    sheetN0 : str
        sheet name
        ex. "ABFarm"
    colidno0 : int
        column number of 'idno0' (sheetN0 )
    wbN1 : str
        Excelfile list of Farm cow's member
        ex. "AB_cowslist.xlsx"
    sheetN1 : str
        sheet name
        ex. "cowslist"
    colidno1 : int
        column number of 'idno1' (sheetN1 new data)

    Returns
    -------
    None.

    """
    #import chghistory
    #import fmstls
    
    wb0obj = chghistory.fpyopenxl(wbN0, sheetN0)
    #wb0 = wb0obj[0]
    sheet0 = wb0obj[1]
    max_row0 = sheet0.max_row
    
    wb1obj = chghistory.fpyopenxl(wbN1, sheetN1)
    wb1 = wb1obj[0]
    sheet1 = wb1obj[1]
    max_row1 = sheet1.max_row
   
    for row1 in range(2, max_row1+1):
        idno1 = fmstls.fpygetCell_value(sheet1, row1, colidno1) 
        
        for row0 in range(2, max_row0+1):
            idno0 = fmstls.fpygetCell_value(sheet0, row0, colidno0)
            if idno1 == idno0:
                xllists = chghistory.fpyxllist_to_indlist_s( sheet0, 11, idno0 )
                #print('xllixts')
                #rint(xllists)

                #breed
                breed = xllists[0][5]
                if breed == "ホルスタイン種":
                    breed = "H"
                elif breed == "黒毛和種":
                    
                    breed = "W"
                else:
                    breed = "unknown"
            
                fmstls.fpyifNone_inputCell_value(sheet1, row1, 6, breed)
        
                #birthday
                birthday = xllists[0][2]
                fmstls.fpyifNone_inputCell_value(sheet1, row1, 7, birthday)
        
                #sex
                sex = xllists[0][3]
                if sex == "メス":
                    sex = "f"
                elif sex == "オス":
                    sex = "m"
                else:
                    sex = "unknown"
        
                fmstls.fpyifNone_inputCell_value(sheet1, row1, 8, sex)
        
                #damidNo
                damidNo = xllists[0][4]
                fmstls.fpyifNone_inputCell_value(sheet1, row1, 11, damidNo)
                
            else:
                continue
    
    wb1.save(wbN1)

#fpyind_trsinf_to_cowslist##################################################
"""
fpyind_trsinf_to_cowslist:
    input individual transfer informations of cowshistory's data
    to cowslist
    v1.0
    2022/8/5
    '検索年月日' -> fillin_date を追加　#*
    v1.01
    2024/1/5
    @author: jicc
    
    
"""

def fpyind_trsinf_to_cowslist(wbN0, sheetN0, 
                              colidno0, wbN1, sheetN1, colidno1, name):
    """
    input individual transfer informations of cowshistory's data
    to cowslist    

    Parameters
    ----------
    wbN0 : str
        Excelfile  data of transfer information
        ex. "AB_cowhistory.xlsx"
    sheetN0 : str
        sheet name
        ex. "ABFarm"
    colidno0 : int
        column number of 'idno0' (sheetN0 )
    wbN1 : str
        Excelfile list of Farm cow's member
        ex. "AB_cowslist.xlsx"
    sheetN1 : str
        sheet name
        ex. "cowslist"
    colidno1 : int
        column number of 'idno1' (sheetN1 new data)
    name : str
        Farm's name etc. "氏名または名称"

    Returns
    -------
    None.

    """
    #import chghistory
    #import fmstls
    
    wb0obj = chghistory.fpyopenxl(wbN0, sheetN0)
    #wb0 = wb0obj[0]
    sheet0 = wb0obj[1]
    max_row0 = sheet0.max_row
    
    wb1obj = chghistory.fpyopenxl(wbN1, sheetN1)
    wb1 = wb1obj[0]
    sheet1 = wb1obj[1]
    max_row1 = sheet1.max_row
   
    for row1 in range(2, max_row1+1):
        idno1 = fmstls.fpygetCell_value(sheet1, row1, colidno1) 
        
        for row0 in range(2, max_row0+1):
            idno0 = fmstls.fpygetCell_value(sheet0, row0, colidno0)
            if idno1 == idno0:
                xllists = chghistory.fpyxllist_to_indlist_s( sheet0, 12, idno0 )
                xllists.sort(key = lambda x:x[8]) 
                #lists' listを 異動年月日 昇順 でsort lambda関数を利用
                #print('xllixts')
                #print(xllists)
                l = len(xllists)
                #l0 = len(xllists[0])
                for i in range(0, l):
                    if xllists[i][10] == name:  #氏名または名称
                        if xllists[i][7] == "出生" : #異動内容
                            #異動年月日->in_date
                            fmstls.fpyinputCell_value(sheet1, row1, 13, xllists[i][8])
                            #異動内容->in_reason
                            fmstls.fpyinputCell_value(sheet1, row1, 14, xllists[i][7])
                            #氏名または名称->from
                            fmstls.fpyinputCell_value(sheet1, row1, 15, xllists[i][10])
                        elif xllists[i][7] == "転入" : #異動内容
                            #異動年月日->in_date
                            fmstls.fpyinputCell_value(sheet1, row1, 13, xllists[i][8])
                            #異動内容->in_reason
                            fmstls.fpyinputCell_value(sheet1, row1, 14, xllists[i][7])
                            if i > 0:
                                #氏名または名称->from
                                fmstls.fpyinputCell_value(sheet1, row1, 15, xllists[i-1][10])
                            else:
                                continue
                            
                            #out_date -> ''  転入した時点で out information clear
                            fmstls.fpyinputCell_value(sheet1, row1, 16, '')
                            #out_reason -> ''
                            fmstls.fpyinputCell_value(sheet1, row1, 17, '')
                            #to -> ''
                            fmstls.fpyinputCell_value(sheet1, row1, 18, '')
                            
                        elif xllists[i][7] == "転出" : #異動内容 
                            #異動年月日->out_date
                            fmstls.fpyinputCell_value(sheet1, row1, 16, xllists[i][8])
                            #異動内容->out_reason
                            fmstls.fpyinputCell_value(sheet1, row1, 17, xllists[i][7])
                            #氏名または名称->to
                            if i < l-1:
                                fmstls.fpyinputCell_value(sheet1, row1, 18, xllists[i+1][10])
                            else:
                                continue
                        elif xllists[i][7] == "死亡" : #異動内容 
                            #異動年月日->out_date
                            fmstls.fpyinputCell_value(sheet1, row1, 16, xllists[i][8])
                            #異動内容->out_reason
                            fmstls.fpyinputCell_value(sheet1, row1, 17, xllists[i][7])
                            #氏名または名称->to
                            fmstls.fpyinputCell_value(sheet1, row1, 18, xllists[i][10])
                        elif xllists[i][7] == "搬入" or "と畜" or "取引": #異動内容 
                            #異動年月日->out_date
                            fmstls.fpyinputCell_value(sheet1, row1, 16, xllists[i][8])
                            #異動内容->out_reason
                            fmstls.fpyinputCell_value(sheet1, row1, 17, xllists[i][7])
                            #氏名または名称->to
                            fmstls.fpyinputCell_value(sheet1, row1, 18, xllists[i][10])
                    
                    #'検索年月日' -> fillin_date
                    fmstls.fpyinputCell_value(sheet1, row1, 20, xllists[i][11])  #*
                    
    wb1.save(wbN1)

#fpysepclst_outfrmin#############################################################
"""
fpysepclst_outfrmin
    separate move-out cows from move-in in a cowslist
    cowslistのExcelfile: AB_cowslist.xlsx の　sheet　ABFarmの情報を
    基準日(最終検索日)における所属牛（転入牛move-in)と転出牛(move-out)の情報に分け、
    2枚のsheet ABFarmin, ABFarmout を作成する
    注) 使用前に２枚のsheet sheetN+'in'と sheetN+'out'を作成しておくこと
        chghistory.fpymkxlsheet(wbN, sheetN, scolN, r)
    v1.0
    2024/1/19
    @author: jicc
    
"""
def fpysepclst_outfrmin( wbN, sheetN, ncol, index ):
    """
    separate move-out cows from move-in in a cowslist
    

    Parameters
    ----------
    wbN : str
        Excelfile to check move-in or move-out data  
        'AB_cowslist.xlsx'　対象のエクセルファイル名
    sheetN : str
        sheet name to separate move-out cows from move-in
        'ABFarm'　対象のエクセルシート名
    ncol : int
        number of columns sheet ABFarm のリストの列数 : 20
    index : int
        index number of an element(out_date)　
        リスト上の　'out_date'のindex番号

    Returns
    -------
    None.

    """

    #import openpyxl
    #import chghistory
    #import datetime

    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    sheetin = wb[sheetN + 'in']
    sheetout = wb[sheetN + 'out']
    
    xllists = chghistory.fpyxllist_to_list_s(sheet, ncol)
    print('xllists')
    print(xllists[0])
    print(xllists[1])
    print(xllists[2])
    
    farmin = [] #default 所属牛（転入牛move-in)
    farmout = [] #default 転出牛(move-out)
    
    lxllists = len(xllists)
    for i in range(0,lxllists):
        
        out_date = xllists[i][index] #out_date
        if type(out_date) == datetime.datetime:
            farmout.append(xllists[i])
        elif out_date == None:
            farmin.append(xllists[i])
        else:
            xllists[i][18] = "check the value of column 16(out_date)"
            #index 18 : note
            farmin.append(xllists[i])

    #print('farmout')
    #for j in range(0,3):
    #    print(farmout[j])
    
    #print('farmin')
    #for k in range(0,3):
    #    print(farmin[k])
    
    lfarmin = len(farmin)
    for j in range(0,lfarmin):
        
        chghistory.fpylisttoxls_s(farmin[j], 1, sheetin)
    
    lfarmout = len(farmout)
    for k in range(0,lfarmout):
        
        chghistory.fpylisttoxls_s(farmout[k], 1, sheetout)
        
                
    wb.save(wbN)    

#fpymod_cowslist############################################################
"""
fpymod_cowslist :  modify cowslist data from a new cowslist
    
    v1.0
    2024/2/11
    @author: jicc
    
"""
def fpymod_cowslist( wbN, sheetN, snewN, ncol ):
    """
    modify cowslist data from a new cowslist

    Parameters
    ----------
    wbN : str
        Excelfile to check move-in or move-out data  
        'AB_cowslist.xlsx'　対象のエクセルファイル名
    sheetN : str
        sheet name of an original cowslist(registered cows)
        'cowslist', 'cowslist2024' etc 追加する対象のエクセルシート名
    snewN : str
        sheet name of a new cowslist
        'cowslistyyyymmdd'
    ncol : int
        number of columns sheet cowslist のリストの列数 : 20

    Returns
    -------
    None

    """
    #import openpyxl
    #import chghistory
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    snew = wb[snewN]
    
    #sheet と　snew の Excel sheetを　それぞれ listにする
    xllists = chghistory.fpyxllist_to_list_s(sheet, ncol)
    xllists_new = chghistory.fpyxllist_to_list_s(snew, ncol)
    
    #リストのエレメント数 : the number of elements of each list
    lxllists = len(xllists)
    lxllists_new = len(xllists_new)
    
    for i in range(0, lxllists):
        for j in range(0,lxllists_new):
            #original idNo == new cows' idNo
            if xllists[i][1] == xllists_new[j][1]:
                #make a comparison between registered and new cows' eack element
                for k in range(2,12):
                    if xllists[i][k] != xllists_new[j][k]:
                        #overwrite a new element
                        xllists[i][k] = xllists_new[j][k]
                    else:
                        continue
            else:
                continue
    
    #overwrite a registered cows' sheet
    chghistory.fpylisttoxls_s_ow(xllists, 1, sheet)
    
    wb.save(wbN)  

#fpyreg_newcows##############################################################
"""
fpyreg_newcows : register new cows from a new cowslist
    
    v1.0
    2024/2/11
    @author: jicc
    
"""
def fpyreg_newcows( wbN, sheetN, snewN, ncol ):
    """
    register new cows from a new cowslist

    Parameters
    ----------
    wbN : str
        Excelfile to check move-in or move-out data  
        'AB_cowslist.xlsx'　対象のエクセルファイル名
    sheetN : str
        sheet name of an original cowslist
        'cowslist', 'cowslist2024' etc 追加する対象のエクセルシート名
    snewN : str
        sheet name of a new cowslist
        'cowslistyyyymmdd'
    ncol : int
        number of columns sheet cowslist のリストの列数 : 20

    Returns
    -------
    None

    """
    #import openpyxl
    #import chghistory
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    snew = wb[snewN]
    
    #sheet と　snew の Excel sheetを　それぞれ listにする
    xllists = chghistory.fpyxllist_to_list_s(sheet, ncol)
    xllists_new = chghistory.fpyxllist_to_list_s(snew, ncol)
    
    #リストのエレメント数 : the number of each list
    lxllists = len(xllists)
    lxllists_new = len(xllists_new)
    
    #振り分け用のlistの初期化　: 
    xllists0 = []  #new cows list default
    for h in range(0,lxllists_new):
        xllists0.append(xllists_new[h])
    #xllists0 = xllists_new とすると、
    #xllists0, xllists_new ともにremoveされ　index errorが起こる 2024/2/9
    
    xllists1 = []   #registered cows list default
    
    #xllists01 = [xllists0, xllists1]
    
    #登録牛と新規牛を分ける
    #separate new cows from registered cows
    for i in range(0, lxllists):
        for j in range(0,lxllists_new):
            if xllists[i][1] == xllists_new[j][1]: #registered
                #the idNo of xllixts[i]  equal th idNo of xllists_new[j]
                xllists1.append(xllists_new[j]) #未使用
                xllists0.remove(xllists_new[j])
                break
            else:
                continue
    
    print('xllists0')
    l0 = len(xllists0)
    print(l0)
    print(xllists0)
    print('xllists1')
    l1 = len(xllists1)
    print(l1)
    xllists1[1]
    
    #新規牛のcowlist_id を修正する : modify cowlist_id of xllists0[k][0]
    l = lxllists+1
    for k in range(0,l0):
        xllists0[k][0] = l
        l = l + 1
    
    #新規牛をExcelsheet cowslist追加 : add new cows to a cowslist
    chghistory.fpylisttoxls_s_(xllists0, 1, sheet)
    
    wb.save(wbN) 

#fpyext_cwslst_at_base_date####################################################
"""
fpyext_cwslst_at_base_date
    extract a cowslist at a base_date in an original cowslist
    original cowslist から　基準日のcowslistを抽出して、
    基準日の cowslist sheet cowslistyyyy_mm_dd　を作成する
    v1.0
    2024/2/16
    基準日の cowslist sheet cowslistyyyymmdd　を作成するように変更 #*
    v1.01
    2024/2/25
    @author: jicc
    
"""
def fpyext_cwslst_at_base_date( wbN0, sN0, coln0, ncol0, index, name, 
                    wbN1, sN1, coln1, ncol1, bdate ):
    """
    extract a cowslist at a base_date in an original cowslist
    original cowslist から　基準日のcowslistを抽出する
    基準日の cowslist sheet cowslistyyyy_mm_dd　を作成する

    Parameters
    ----------
    wbN0 : str
        Excelfile to check move-in or move-out data  
        'AB_cowshistory.xlsx'　異動情報収納ファイル
    sN0 : str
        sheet name of cows' history data 
        'ABFarm'　異動情報をリストしたシート名
    coln0 : int
        column's number of idNo in sheet sN0 
        sheet sN0 の個体識別番号の入っている列番号
    ncol0 : int
        number of columns of sheet sN0's list
        sheet sN0 のリストの列数
    index : int
        index number of an element(Farm name)　
        リスト上の　'氏名または名称'のindex番号
    name : str
        Farm name '氏名または名称'
    wbN1 : str
        Excelfile of a farm's cowslist
        'AB_cowslist.xlsx'　cowslist 個体情報収納ファイル 
    sN1 : str
        sheet name of cowslist
        'cowslist' , 'cowslist2024' etc　牧場の所属個体情報をリストしたシート名 
    coln1 : int
        column's number of idNo in sheet sN1 
        sheet sN1 の個体識別番号の入っている列番号
    ncol1 : int
        number of columns of sheet sN1's list
        sheet sN1 のリストの列数
    bdate : str
        base date 基準日

    Returns
    -------
    None.

    """
    import openpyxl
    import chghistory
    import datetime
    import fmstls

    wb0 = openpyxl.load_workbook(wbN0) 
    s0 = wb0[sN0]
    
    wb1 = openpyxl.load_workbook(wbN1)
    s1 = wb1[sN1]
    
    #bdate_ = chghistory.fpyreplace_str(bdate, '/', '_')
    bdate_ = fmstls.fpystrdate_to_yyyymmdd( bdate )   #*
    #print(bdate)
    #print(bdate_)
    sN2 = 'cowslist' + bdate_
    #sheet名 sN2 のsheetがなければ作成する
    s2 = chghistory.fpymkxlsheet_(wb1, sN2, 'columns', 1)
    
    #wb1.save(wbN1)
    
    #AB_cowslist.xlsx/cowslist(org)からリストを作成する
    xllist1 = chghistory.fpyxllist_to_list_s(s1, ncol1)
    
    lxllist1 = len(xllist1)
    
    for i in range(0, lxllist1):
        
        xllists0 = chghistory.fpyxllist_to_indlist_s(s0, ncol0, xllist1[i][coln1-1])
        #個体識別番号 idNo の異動情報のリスト
        
        if xllists0 == []:
            idNotmp = xllist1[i][coln1-1]
            print("xllists0")
            print(xllists0)
            print('cow ' + idNotmp + 'には、異動情報がありません。')
        
        else:
        
            #lxllists0 = len(xllists0)
            xllists0.sort(key = lambda x:x[8]) #, reverse=True
            #lists' listを 異動年月日 昇順 でsort lambda関数を利用
            #No ([6])昇順でsortしたほうが良いかもしれない。2024/1/13
        
            xllists0_ =chghistory.fpyext_frmlsts_lst(xllists0, index, name)
            #index 10 : "氏名または名称"
            #当該牧場の異動情報だけ抽出

            xllists0_.sort(key = lambda x:x[8]) #, reverse=True
            #lists' listを 異動年月日 昇順 でsort lambda関数を利用
        
            #print("xllists0_")
            #print(xllists0_)
        
            xllists0_ = chghistory.fpyarr_frmlsts_lst( xllists0, xllists0_ )   # *)
            #最後の"転出"が欠けていた場合の調整
            #print("xllists0_")
            #print(xllists0_)
        
            terms_in_farm = chghistory.fpyterms_in_farm_( xllists0_ )
            #当該牧場にいた滞在期間
            #print("terms_in_farm")
            #print(terms_in_farm)
        
            #bdate = '2023/12/31'
            print(bdate)
            #基準日
            if type(bdate) == str: 
                bdate = datetime.datetime.strptime(bdate, '%Y/%m/%d')
                #datetimeに変換
                #print('bdate')
                #print(bdate)
        
            belongornot = chghistory.fpyind_belongornot( bdate, terms_in_farm )
            #基準日にその農場に所属していたかどうか
            #print('belongornot')
            #print(belongornot)
        
            if belongornot == 0: #move-out
                continue
        
            elif belongornot == 1: #move-in belonging
            
                chghistory.fpylisttoxls_s(xllist1[i], 1, s2)
            
                  
    wb1.save(wbN1)

#fpycowslistManual###########################################################
'''
    AB_cowslist.xlsx 作成のための　PS用マニュアル
    v1.0 by jicc
    2022/7/31 by jicc

Umotionの　飼養牛一覧yyyymmdd から
    cowslistyyyymmdd を作成する

'''    
def fpycowslistManual():
    
    print('-----cowslist Manual-----------------------------------------------------v1.02------')
    print('1. Umotion 飼養牛一覧　父牛掲載yyyymmdd.csv を')
    print(' AB_cowslist.xlsx/sheet\"cowslistyyyymmddorg\"として移行する')
    print(' ')
    print('2.データ移行用の列名だけのsheet\"cowslistyyyymmdd\"を作成する')
    print('  python ps_fpynewsheet_args.py wbN sheetN scolN row')
    print('wbN : AB_cowslist.xlsx, sheetN : cowslistyyyymmdd, scolN : columns, row : 1')
    print(' ')
    print('3.sheet \"cowlistyyyymmdd\" にデータ入力する')
    print('PS> python ps_fpydf_cowslist_args.py wbN sheetorg sheetN fillinDate')
    print('  wbN : AB_cowslist.xlsx, sheetorg : cowslistyyyymmddorg,')
    print(' sheetN : cowslistyyyymmdd, fillinDate : yyyy/mm/dd')
    print(' ')
    print('4.sheet cowslistyyyymmdd 2列 cowidNoを10桁文字列に統一する')
    print(' PS> python ps_fpyidno_9to10_args.py wbN sheetN col')
    print(' wbN : AB_cowslist.xlsx, sheetN : cowslistyyyymmdd, col : 2')    
    print(' ')
    print('5.input individual informations from cowshistory\'s data to cowslist')
    print(' PS> python ps_fpyind_inf_to_cowslist_args.py wbN0 sheetN0 colidno0 wbN1 sheetN1 colidno1')
    print(' wbN0 : AB_cowshistory.xlsx, sheetN0 : ABFarm, colidno0 : 2 (column number fo idno0), ') 
    print(' wbN1 : AB_cowslist.xlsx, sheetN1 : cowslist, colidno1 : 2 (column number fo idno1)')
    print(' ')
    print('6.input individual transfer informations from cowshistory\'s data to cowslist')
    print(' PS> python ps_fpyind_trsinf_to_cowslist_args.py wbN0 sheetN0 colidno0')
    print(' wbN1 sheetN1 colidno1 name')
    print(' wbN0 : AB_cowshistory.xlsx, sheetN0 : ABFarm, colidno0 : 2 (column number fo idno0), ') 
    print(' wbN1 : AB_cowslist.xlsx, sheetN1 : cowslist, colidno1 : 2 (column number fo idno1)')
    print(' name : 氏名または名称')
    print(' ')
    print('---------------------------------------------------------2022/8/5 by jicc---------')
    

#fpycowslistManual00##########################################################    
'''
    fpycowslistManual00:
    AB_cowslist.xlsx 作成のための　PS用マニュアル
    v2.0 by jicc
    2024/1/14 by jicc

    AB_cowshistory.xlsx の個体情報、異動情報から　
    cowslistyyyymmdd を作成する

'''    
def fpycowslistManual00():

    print('-----cowslist Manual00---------------------------------------------------v2.2------')    
    print('1.input individual informations from cowshistory\'s data to cowslist')
    print(' PS> python ps_fpyind_inf_to_cowslist_args.py wbN0 sheetN0 colidno0 wbN1 sheetN1 colidno1')
    print(' wbN0 : AB_cowshistory.xlsx, sheetN0 : ABFarm, colidno0 : 2 (column number fo idno0), ') 
    print(' wbN1 : AB_cowslist.xlsx, sheetN1 : cowslist, colidno1 : 2 (column number fo idno1)')
    print(' ')
    print('2.input individual transfer informations from cowshistory\'s data to cowslist')
    print(' PS> python ps_fpyind_trsinf_to_cowslist_args.py wbN0 sheetN0 colidno0')
    print(' wbN1 sheetN1 colidno1 name')
    print(' wbN0 : AB_cowshistory.xlsx, sheetN0 : ABFarm, colidno0 : 2 (column number fo idno0), ') 
    print(' wbN1 : AB_cowslist.xlsx, sheetN1 : cowslist, colidno1 : 2 (column number fo idno1)')
    print(' name : 氏名または名称')
    print(' ')
    print('3.modify cowslist data from a new cowslist')
    print(' PS> python ps_fpymod_cowslist_args.py wbN sheetN snewN ncol')
    print(' wbN: ..\AB_cowslist.xlsx, sheetN: cowslist, scolN: cowslist_new, ncol: 20')
    print(' 新しい情報に変更があれば修正する ') 
    print(' ')
    print('4.register new cows from a new cowslist')
    print(' PS> python ps_fpyreg_newcows_args.py wbN sheetN snewN ncol')
    print(' wbN: ..\AB_cowslist.xlsx, sheetN: cowslist, scolN: cowslist_new, ncol: 20')
    print(' 新規牛をExcelsheet cowslist追加 ')
    print(' ')
    print('5.make an ExcelSheet if it dose not exist')
    print(' PS> python ps_fpymkxlsheet_args.py wbN sheetN scolN r')
    print(' wbN: ..\AB_cowslist.xlsx, sheetN: ABFarmout, scolN: columns, r: 1')
    print(' 2枚のsheet ABFarmin, ABFarmout がなければ、作成する ') 
    print(' ')
    print('6.separate move-out cows from move-in in a cowslist')
    print(' PS> python ps_fpysepclst_outfrmin_args.py wbN sheetN ncol index')
    print(' wbN: ..\AB_cowslist.xlsx, sheetN: ABFarm, ncol: 20, r: index : 15')
    print(' 検索日における所属牛（転入牛move-in)と転出牛(move-out)の情報に分け、 ') 
    print(' 2枚のsheet ABFarmin, ABFarmout を作成する ')
    print(' ')
    print('7.extract a cowslist at a base_date in an original cowslist')
    print(' PS>  ps_fpyext_cwslst_at_base_date_args.py wbN0 sN0 coln0 ncol0 index name')
    print(' wbN1 sN1 coln1 ncol1 bdate')
    print(' wbN0 : AB_cowshistory.xlsx, sN0 : ABFarm, coln0 : 2, ncol0 : 12,')
    print(' index : 10, name : \'Farm name\', wbN1 : AB_cowslist.xlsx,')
    print(' sN1 : cowslist2024, coln1 : 2, ncol1 : 20, bdate : 2024/1/14')
    print(' original cowslist から　基準日のcowslistを抽出して、 ') 
    print(' 基準日の cowslist sheet cowslistyyyy_mm_dd　を作成する')
    print('---------------------------------------------------------2024/2/18 by jicc---------')
    
    
    
    
    