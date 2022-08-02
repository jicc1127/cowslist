# -*- coding: utf-8 -*-
import os, re
import tabula
import openpyxl
import csv
import shutil
import datetime
####################################################from fmstls.py##########
"""
fpyopenxl(wbN, sheetN):
    Excelfile wbN.xlsx　sheet sheetN Open 
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopenxl(wbN, sheetN):
    """
    Excelfile wbN.xlsx　sheet sheetN Open

    Parameters
    ----------
    wbN : str
        ExcelFile Name   ex.MH_CowHistory.xlsx
    sheetN : str
        sheet name

    Returns
    -------
    None.

    """
    
    #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    return [wb, sheet]
"""
fpyopencsv_robj:
    csvfile Open for Reader object
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopencsv_robj(csvN):
    """
    csvfile Open for Reader object

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    None.

    """
    #import csv
    
    #filename = csvN.split('.')
    #filename = filename[0]  #拡張子を削除したfilename
    
    filename_file = open(csvN)     #csvfile open
    filename_reader = csv.reader(filename_file)       #get Reader object
    
    
    return filename_reader
    
"""
fpyopencsv_rdata:
    csvfile Open for Reader data
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopencsv_rdata(csvN):
    """
    csvfile Open for Reader data

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    None.

    """
    #import csv
    
    #filename = csvN.split('.')
    #filename = filename[0]  #拡張子を削除したfilename
    
    filename_file = open(csvN)     #csvfile open
    filename_reader = csv.reader(filename_file)       #get Reader object
    filename_data = list(filename_reader)             #list's list
    
    return filename_data
    
"""
fpyopencsv_w:
    csvfile Open for Writer
    v1.00
    2022/1/7
    @author: jicc
"""
def fpyopencsv_w(csvN):
    """
    csvfile Open for Writer

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    None.

    """
    #import csv
    
    output_file = open(csvN, 'w', newline='')       #csvfile open
    output_writer = csv.writer(output_file)       #get Reader object
     
    return output_writer

"""
fpygetCell_value: get value from the target Cell
v1.00
2022/2/4

@author: inoue
"""
def fpygetCell_value(sheet, r, col):
    """
    get value from the target Cell on an Excelsheet

    Parameters
    ----------
    sheet : worksheet
        sheetBLV
    r : int
        row
    col : int
        column

    Returns
    -------
    value

    """

    value = sheet.cell(row=r, column=col).value
    return value

"""
fpyinputCell_value: input value to the target Cell
v1.00
2022/2/4

@author: inoue
"""
def fpyinputCell_value(sheet, r, col, vl):
    """
    input value to the target Cell

    Parameters
    ----------
    sheet : worksheet
        sheetBLV
    r : int
        row
    col : int
        column
    vl : type of value
    
    Returns
    -------
    None.

    """

    sheet.cell(row=r, column=col).value = vl 

#fpyNewSheet#
"""
fpyNewSheet : Excelbookに
sheet　'columns'と同じ sheet　'scolN'を作成する。
ｖ1.01
2022/5/3

@author: jicc

"""
def fpyNewSheet(wbN, sheetN, scolN, r):
    """
    Excelbookに sheet 'scolN' r行目の'columns'を1行目に配置した sheet'sheetN'を作成する。
    *sheet 'columns'(列名を記入したシート) を作成しておく
    Parameters
    ----------
    wbN : 　str          
        sheetを作成するワークブック
    sheetN : str　　　　　　シート名:"????" 
        作成するシート
    scolN : str         シート名: "columns"
        参照するシート
	r : int		r行目 作成するcolumn行
    Returns
    -------
    None.

    """
    #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    #sheetN = wb[sheetN]
    wb.create_sheet(title=sheetN, index=0)
    sheet = wb[sheetN]
    scol = wb[scolN]
    
    maxcol = scol.max_column #sheet columnの最終列
    
    for i in range(1,maxcol+1):
        sheet.cell(row=r, column=i).value = scol.cell(row=1, column=i).value
    
     
    wb.save(wbN)
    
"""
fpychgSheetTitle      :change ExcelSheet's title
v1.0
2022/3/30

@author: inoue
"""
def fpychgSheetTitle(wbN, sheetN, sheetN1):
    """
    change the sheet's title

    Parameters
    ----------
    wbN : str
        Excelfile to check double data  '??_CowsHistory.xlsx'
    sheetN : str
        元のシート名  : 'KTFarm'
    sheetN1 : str
        変更名      : 'KTFarmorg' 

    Returns
    -------
    None.

    """
    #import chghistory
    wbobj = fpyopenxl(wbN, sheetN)
    wb = wbobj[0]
    sheet = wbobj[1]
    sheet.title = sheetN1
    wb.save(wbN)
 
####################################################from fmstls.py##########
"""
fpypdf_to_csv
    convert ****.pdf to ****.csv file
v1.00
2022/1/1
@author: inoue
"""

#import tabula

def fpypdf_to_csv(filename, Path):
    '''
    filename: string
    csv変換するpdf  filename

    Returns
    -------
    filename.csv

    '''
    filename_pdf = Path + "\\" +  filename + ".pdf"
    filename_csv = Path + "\\" + filename + ".csv"
    tabula.convert_into(filename_pdf, filename_csv, 
                    stream=True , output_format="csv", pages="all")


"""
fpySpdf_in_Dir  -  ディレクトリ内の特定の拡張子を持つファイルを見つけ
                  file名をプリントする
v1.00
2022/1/1
by jicc
"""
def fpySpdf_in_Dir(Ext, Path):
    
    #import os, re
    
    fs = os.listdir(Path) 
    #Pathに指定したフォルダー内の、ファイル名とフォルダー名のリストを返す
    regex_ext = re.compile(Ext)   #Regex(regular expression)オブジェクトを返す
    #print(regex_ext)
    
    for f in fs:
        #print(f)
        mo = regex_ext.search(f) #regx_extにマッチするとmatdhオブジェクトを返す
        if mo: #!=None
            print(f)

######################################################未使用
"""
fpySpdf_in_Dir_to_csv  -  ディレクトリ内の特定の拡張子(.pdf)を持つファイルを見つけ
                  csvfile に変換する
                  変換したpdffileを　フォルダーpdforgに移動する
v1.01
2022/1/11
by jicc

"""
def fpySpdf_in_Dir_to_csv(Ext, Path, bckPath):
    """
    Parameters
    ----------
    Ext : str
        拡張子　　　'\.pdf'　　
    Path : str
        path      '.\\' カレントディレクトリ
    bckPath : str
        file移動するフォルダーのpath

    Returns
    -------
    None.

    """
    
    #import os, re
    #import chghistory
    #import shutil
    fs = os.listdir(Path)
    regex_ext = re.compile(Ext)
    #print(regex_ext)
    
    for f in fs:
        #print(f)
        mo = regex_ext.search(f)
        if mo:
            print(f)
            filename = f.split('.')
            filename = filename[0]
            fpypdf_to_csv(filename, Path)
            #print(mo.group())
            filename_pdf = filename + '.pdf'
            shutil.move(filename_pdf, bckPath)
            
#fpyCowHistory##############################################################            
"""
fpyCowHistory
    牛の個体情報.csvから、CowHistory.csv(changehistory's list )を作成する
    'No'を行頭にいどうすることを中止、単純に個体データと異動データを結合するように変更
    個体識別番号　９桁->１０桁
    日付データ　yyyy.mm.dd -> yyyy/mm/dd の処置を追加
ｖ1.02
2022/1/9
@author: inoue
"""

def fpyCowHistory(csvorgN, csvoutN):
    '''
    牛の個体情報.csvから、CowHistory.csv(changehistory's list )を作成する

    Parameters

    ----------
    csvorgN : str
        もととなるcsvファイル        MH_???_yyyymmdd.csv
    csvoutN : str
        作成するcsvファイル　　　　　　MH_???_yyyymmddH.csv　

    Returns
    -------
    None.

    '''
    #import csv
    #import chghistory
    
    mhcow_file = open(csvorgN)     
    ################################################################
    #csvfile open , encoding="utf-8",  "shift-jis"
    #UnicodeDecodeError: 'cp932' ... のエラーのためencoding="utf-8"を追加でもダメ2022/1/11
    #PS> ps_fpyymd_csvtocowshistory_csv_args.py .csv .\ .\csvorg で実施の時のみ。
    #アナコンダ　インタラクティブシェルで行ったら問題なし。
    #今日試したらError出なかった。　なぜ？？　2022/1/12
    #################################################################
    mhcow_reader = csv.reader(mhcow_file)       #get Reader object
    mhcow_data = list(mhcow_reader)             #list's list
    
    cowhistory_header = mhcow_data[0]
    print(cowhistory_header)
    cowhistory_header = cowhistory_header \
        + ['No', '異動内容', '異動年月日', '飼養施設所在地', '氏名または名称'] #見出し行のリスト
     
    output_file = open(csvoutN, 'w', newline='')
    output_writer = csv.writer(output_file)
    output_writer.writerow(cowhistory_header)
    row_max = mhcow_reader.line_num  # =len(mhcow_data) リストの行数
    
    id_info = mhcow_data[1]  	
    #['個体識別番号', '出生の年月日', '雌雄の別', '母牛の個体識別番号', '種別']
    #1行目の絶対データ
    #print(id_info)
    #id_info_ = id_info[0]
    id_info[0] = fpycsvidNo_9to10( id_info[0] ) #idNo
    id_info[1] = fpydate_dottoslash( id_info[1]) #出生年月日 'yyyy.mm.dd' -> 'yyyy/mm/dd'
    id_info[3] = fpycsvidNo_9to10( id_info[3] ) #damidNo
    print(id_info)
    for row_num in range(5, row_max):
  
        history = mhcow_data[row_num] 	
        #['No', '異動内容', '異動年月日', '住所', '氏名または名称']
        #row_num行目の相対データ
        #print(history)
        history[2] = fpydate_dottoslash( history[2]) #異動年月日 'yyyy.mm.dd' -> 'yyyy/mm/dd'
        print(history)
        id_info_history = id_info + history #行データを結合
        output_writer.writerow(id_info_history)
    
    output_file.close()

#fpycsvlisttoxls############################################################            
"""
fpycsvlisttoxls: 
    csvfileのデータをexcelfileに移行する
    死亡のテーブルを回避する処置を加えた
    ｖ1.01
    2022/1/13　
    @author: jicc
    
"""
def fpycsvlisttoxls(csvN, wbN, sheetN):
    """
    csvfileのデータをexcelfileに移行する

    Parameters
    ----------
    csvN : str
        original csvfile  'MH_???_History.csv'
    wbN : str
        Excelfile to move History data  'MH_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   'MHFarm' 

    Returns
    -------
    None.

    """
    #import chghistory
    #import openpyxl
    
    wbobj = fpyopenxl(wbN, sheetN)   #get Worksheet object
    wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    max_row = sheet.max_row                     
    
    csvdata = fpyopencsv_rdata(csvN)     #list's list of the csvfile
    ln = len(csvdata)       #the length of the list csvdata
    ln_ = len(csvdata[0])   #the number of the csvdata's list[0]
    k = max_row
    for i in range(1, ln):
        No = csvdata[i][5]  #死亡のとき、'異動内容「死亡」の'以下のデータを削除のため
        #* "1", "2", "異動内容「死亡」の"　strlength <=2 で振り分け　2022/1/13 ｖ1.01
        No_ln = len(No)      
        while (No_ln<=2):       #* #No==99まで可能
            for j in range(0, ln_+1):
                if j== 0:
                    sheet.cell(row=max_row+i, column=j+1).value = k
                    #print(k)
                    k = k + 1
                else:
                    sheet.cell(row=max_row+i, column=j+1).value = \
                        csvdata[i][j-1]
                        #l = csvdata[i][j-1]
                        #print(l)
            break #*
            
        
    wb.save(wbN)

#fpycsvidNo_9to10###########################################################    
"""
fpycsvidNo_9to10:
    idNo in a csvfile 9figures to 10figures
    v1.00
    2022/1/9
    @author: inoue
    
"""
def fpycsvidNo_9to10( idNo ):
    """
    idNo in a csvfile 9figures to 10figures

    Parameters
    ----------
    idNo : str
        idNo

    Returns
    -------
    None.

    """
    if len(idNo) == 9:
        idNo = '0' + idNo 
    else:
        idNo = idNo 
    
    return idNo 

"""
fpydate_dottoslash:
    date in a csvfile 'yyyy.mm.dd' to 'yyyy/mm/dd'
    v1.02
    2022/7/6
    @author: inoue
    
"""
def fpydate_dottoslash( date ):
    """
    date in a csvfile 'yyyy.mm.dd' to 'yyyy/mm/dd'

    Parameters
    ----------
    date : str
        date

    Returns
    -------
    date  : datetime 

    """
    date = date.split('.')
    date = "/".join(date)   #*
    #date = datetime.datetime.strptime(date, '%Y/%m/%d')
    #date(str)をdatetimeに変換　ｖ1.01　2022/3/1
    #date = date.strftime('%Y/%m/%d')
    #'yyyy/mm/dd'に変換 v1.02 2022/7/6 #*の状態と同じ解消する2022/7/7
    return date
 
"""
fpyymd_csvtoCowsHistory_csv:
    フォルダー内の個体履歴org(csv)をCowsHistory.csvに変更する
    変更後orgcsvfile を　別フォルダー(./csvorg)に移動する
    v1.01
    2022/1/10
    @author: jicc
"""
def fpyymd_csvtoCowsHistory_csv(Ext, Path, bckPath):
    """
    Parameters
    ----------
    Ext : str
        拡張子　　　'\.csv'　　
    Path : str
        path      '.\\' カレントディレクトリ
    bckPath : str
        file移動するフォルダーのpath　'.\\csvorg' 

    Returns
    -------
    None.

    """
    
    #import os, re
    #import shutil
    #import chghistory
    fs = os.listdir(Path)
    regex_ext = re.compile(Ext)
    #print(regex_ext)
        
    for f in fs:
        #print(f)
        mo = regex_ext.search(f)
        if mo:
            print(f)
            f_ = f.split('.')
            csvoutN = f_[0] + 'H.csv'
            csvorgN = f
            fpyCowHistory(csvorgN, csvoutN)
            
            try:
                shutil.move(csvorgN, bckPath)
                #csvoriginalfile(csvodgN) を　フォルダーbckPathに移動
            except shutil.Error:
                print( csvorgN + ' already exists') 
     
            
            
"""
fpyHistory_csvto_xlsx:
    フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する
    移動後???H.csvを　別フォルダー(./csvhistory)に移動する
    v1.02
    2022/7/5
    @author: jicc
"""
def fpyHistory_csvto_xlsx(Ext, Path, bckPath, wbN, sheetN):
    """
    フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する
    移動後???H.csvを　別フォルダー(./csvhistory)に移動する
    add try~except v1.02 2022/7/5
    Parameters
    ----------
    Ext : str
        拡張子　　　'\.csv'　　
    Path : str
        path      '.\\' カレントディレクトリ
    bckPath : str
        file移動するフォルダーのpath
    wbN : str
        Excelfile to move History data  'MH_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   'MHFarm' 

    Returns
    -------
    None.

    """
    
    #import os, re
    #import shutil
    #import chghistory
    fs = os.listdir(Path)
    regex_ext = re.compile(Ext)
    #print(regex_ext)
        
    for f in fs:
        #print(f)
        mo = regex_ext.search(f)
        if mo:
            print(f)
            fpycsvlisttoxls(f, wbN, sheetN)
            
            try:
                shutil.move(f, bckPath)
            except shutil.Error:
                print( f + ' already exists')
            #csvoriginalfile(csvodgN) を　フォルダーbckPathに移動
            #上書きできないので例外処理


"""
fpystrtodatetime : str'yyyy/mm/dd'をdatetime に変換する

v1.00
2022/3/1

@author: inoue
"""
def fpystrtodatetime( date ):
    """
    str'yyyy/mm/dd'をdatetime に変換する

    Parameters
    ----------
    date : str
       'yyyy/mm/dd'

    Returns
    -------
    date  : datetime

    """
    #import datetime
    date = datetime.datetime.strptime( date, '%Y/%m/%d')
    
    return date

"""
fpyxlstrymdtodatetime : Excel cell 'yyyy/mm/dd'をdatetimeに変換する

v1.01
2022/3/2

@author: inoue
"""
def fpyxlstrymdtodatetime(wbN, sheetN, col):
    """
    Excel cell 'yyyy/mm/dd'をdatetimeに変換する

    Parameters
    ----------
    wbN : str
        書き換えするExcelFile名   :??_CowsHistory.xlsx
    sheetN : str
        書き換えするシート名　　　　：??Farm
    col : int
        書き換えする列

    Returns
    -------
    None.

    """
    #import fmstls 
    #import chghistory
    #import datetime
    
    xl = []
    xl = fpyopenxl(wbN, sheetN)
    wb = xl[0] #workbook
    sheet = xl[1] #worksheet
    
    for i in range(2, sheet.max_row+1):
        
        date = fpygetCell_value(sheet, i, col)
        if type(date) == str: #date = 'str'の場合datetimtに変換1.01
        #if type(date) != datetime.datetime: #これではNoneセルでstopする
            date = fpystrtodatetime( date )
            fpyinputCell_value(sheet, i, col, date)
        else:
            continue
            
        
        
    wb.save(wbN)

#fpyxllist_to_list#########################################################
"""
fpyxllist_to_list: 
    excelfileのリストを　lists'　list にする
    
    ｖ1.00
    2022/3/9
    @author: jicc
    
"""
def fpyxllist_to_list(wbN, sheetN, ncol):
    """
    excelfileのデータをlists'listにする

    Parameters
    ----------
    wbN : str
        Excelfile to move History data  '??_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   '??Farm' 
    ncol :  int
        number of columns
    Returns
    -------
    xllists : lists' list

    """
    #import chghistory
    #import openpyxl
    
    wbobj = fpyopenxl(wbN, sheetN)   #get Worksheet object
    #wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    max_row = sheet.max_row
    # max_col = sheet.max_col
    #AttributeError: 'Worksheet' object has no attribute 'max_col'
    xllist = []
    xllists = []
    for i in range(2, max_row+1):  #タイトル行は飛ばす
        
        for j in range(1,ncol+1):
            coldata = sheet.cell(row=i, column=j).value
            xllist.append(coldata)
            
        xllists.append(xllist)
        xllist = []    
    return xllists

#fpyxllsit_to_indlist######################################################
"""
fpyxllist_to_indlist:
    get an individual lists' list from excelfile's list
        
    ｖ1.00
    2022/7/12
    @author: jicc
    
"""
def fpyxllist_to_indlist(wbN, sheetN, ncol, idno):
    """
    get an individual lists' list from excelfile's list

    Parameters
    ----------
    wbN : str
        Excelfile to move History data  '??_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   '??Farm' 
    ncol :  int
        number of columns
    idno : str
        ex. "0123456789"
    Returns
    -------
    xllists : lists' list

    """
    #from jiccModule import chghistory
    #import chghistory
    #import openpyxl
    
    wbobj = fpyopenxl(wbN, sheetN)   #get Worksheet object
    #wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    max_row = sheet.max_row
    # max_col = sheet.max_col
    #AttributeError: 'Worksheet' object has no attribute 'max_col'
    xllist = []
    xllists = []
    for i in range(2, max_row+1):   #タイトル行は飛ばす
        idno_ = fpygetCell_value(sheet, i, 2) 
        #excellist's idno column 2
        if idno_ == idno:
            for j in range(1,ncol+1):
                coldata = sheet.cell(row=i, column=j).value
                xllist.append(coldata)
            
            xllists.append(xllist)
            xllist = [] 
            
    return xllists

#fpyxllist_to_indlist_s######################################################
"""
fpyxllist_to_indlist_s:
    get an individual lists' list from excelfile's list
    arguments 'wbN, sheetN' -> 'sheet' worksheetobject version    
    ｖ1.00
    2022/7/17
    @author: jicc
    
"""
def fpyxllist_to_indlist_s(sheet, ncol, idno):
    """
    get an individual lists' list from excelfile's list

    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    ncol :  int
        number of columns
    idno : str
        ex. "0123456789"
    Returns
    -------
    xllists : lists' list

    """
    #import chghistory
  
    max_row = sheet.max_row

    xllist = []
    xllists = []
    for i in range(2, max_row+1):   #タイトル行は飛ばす
        idno_ = fpygetCell_value(sheet, i, 2) 
        #excellist's idno column 2
        if idno_ == idno:
            for j in range(1,ncol+1):
                coldata = sheet.cell(row=i, column=j).value
                xllist.append(coldata)
            
            xllists.append(xllist)
            xllist = [] 
            
    return xllists

#fpyaddclm_to_lsts_lst####################################################
"""
fpyaddclm_to_lsts_lst : 
   lists'listに最終カラムを追加する
   
   v1.0
   2022/3/28

@author: inoue
"""
def fpyaddclm_to_lsts_lst(xllists, colv):
    """
    lists'listに最終カラムを追加する

    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile
    colv : int str None etc
        
    Returns
    -------
    最終列を追加した　lists'list 

    """
    
    lxll = len(xllists)
    for i in range(0, lxll):
        xllists[i].append(colv)
    return xllists

#fpydelclm_frm_lsts_lst#################################################
"""
fpydelclm_frm_lsts_lst : 
   lists'listのカラムを削除する
   
   v1.0
   2022/3/28

@author: inoue
"""
def fpydelclm_frm_lsts_lst(xllists, col):
    """
    lists'listのカラムを削除する

    Parameters
    ----------
    xllists : lists'list
        lists'list from ExcelFile
    col : int 
    削除する列番号   
    Returns
    -------
    列を削除した　lists'list 

    """
    
    lxll = len(xllists)
    for i in range(0, lxll):
        del xllists[i][col]
    return xllists

#fpyflag_dblrcd_1#######################################################
"""
fpyflag_dblrcd_1 : flag double record 1
   lists'listの重複リストに　1（重複）でチェックを入れる
   
   v1.01
   2022/4/3

@author: inoue
"""
def fpyflag_dblrcd_1(xllists):
    """
    lists'listの重複リストに　1（重複）でチェックを入れる

    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile

    Returns
    -------
    重複リストに "1"を追加した　lists'list 

    """
    
    lxll = len(xllists)
    #xldblrows = []
    for i in range(0, lxll):
        #print(xllists[i])
        #k=0
        
        for j in range(0, i+1):
            #print(xllists[j])
            if j!= i:
                if xllists[i][1:5] == xllists[j][1:5] and xllists[i][7:10] == xllists[j][7:10]:
                #LinNo と No 以外が一致したら v1.01
                    xllists[i][11] = 1

                else:
                    continue
            else:
                continue
            
    return xllists 

"""
fpyflag_dblrcd_1_ : flag double record 1
   2つのlists'list　listorgとlisttmpを比較し、
   listtmpの重複リストに　1（重複）でチェックを入れる
   v1.0
   2022/7/15
   @author: jicc
   
"""
def fpyflag_dblrcd_1_(xllists, trs_inf):
    """
    2つのlists'list　listorgとlisttmpを比較し、
   listtmpの重複リストに　1（重複）でチェックを入れる

    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile original list
    
    trs_inf : lists'list
        lists'list from web search data  

    Returns
    -------
    重複リストに "1"を追加した　lists'list listtmp 

    """
    lxll = len(xllists)
    ltrs = len(trs_inf)
    
    for i in range(1, ltrs):    #columns' list skip
        for j in range(0, lxll):
            #print(xllists[j])
            if trs_inf[i][0:10] == xllists[j][1:]:
                trs_inf[i][10] = 1
            else:
                continue
            
    return trs_inf

#fpydel_dblrcd##############################################################
"""
fpydel_dblrcd : delete double record
   lists'listの重複リストの一つを削除する
   add argument coln v1.01 2022/7/16
   v1.01
   2022/7/16

@author: inoue
"""
def fpydel_dblrcd(xllists, coln, colv):
    """
    lists'listの重複リストの一つを削除する
    
    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile
    coln : int
        column flag's number 
    colv : int
        0, 1

    Returns
    -------
    重複リストに "1"を追加した　lists'list 

    """
    
    lxll = len(xllists)
    xllists_ = []
    for i in range(0, lxll):
        if xllists[i][coln] == colv:
            xllists_.append(xllists[i])
        else:
            continue

    return xllists_ 

#fpylisttoxls############################################################
"""
fpylisttoxls: 
    listのデータをexcelfileに移行する
    ｖ2.0
    2022/7/28
    @author: jicc
    
"""
def fpylisttoxls(xllist, fstcol, wbN, sheetN):
    """
    listのデータをexcelfileに移行する
    開始行　sheet.max_row + 1
    開始列 fstcol

    Parameters
    ----------
    xllist : str
        list from original csvfile  'MH_???_History.csv'
    fstcol : int
        first volumn number to input data
    wbN : str
        Excelfile to move History data  'MH_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   'MHFarm' 

    Returns
    -------
    None.

    """
    #from jiccModule import chghistory
    #import chghistory
    #import openpyxl
    
    wbobj = fpyopenxl(wbN, sheetN)   #get Worksheet object
    wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    max_row = sheet.max_row                     
    rn = max_row + 1 #first row to input records
    ln = len(xllist)
           #the length of the list xllist
    if ln > 0: #リストに要素がない場合を排除 v1.01 2022/4/3
        ln_ = len(xllist[0])   #the number of the xllist's list[0]
        for i in range(0, ln):
            for j in range(0, ln_):
                sheet.cell(row=rn, column=j+fstcol).value = xllist[i][j]
            rn = rn + 1
            print('add a new transfer informatyon')
    else:
    	print(' xllist have no element!')
        
    wb.save(wbN)        

#fpylisttoxls_s############################################################
"""
fpylisttoxls: 
    listのデータをexcelfileに移行する
    ｖ2.0
    2022/7/28
    @author: jicc
    
"""
def fpylisttoxls_s(xllist, fstcol, sheet):
    """
    listのデータをexcelfileに移行する
    開始行　sheet.max_row + 1
    開始列 fstcol
    arguments 'wbN, sheetN' -> 'sheet' worksheetobject version 
    
    Parameters
    ----------
    xllist : str
        list from original csvfile  'MH_???_History.csv'
    fstcol : int
        first volumn number to input data
   sheet : worksheet.worksheet.Worksheet
        worksheet object

    Returns
    -------
    None.

    """
    #import chghistory
    #import openpyxl
    
    max_row = sheet.max_row                     
    rn = max_row + 1 #first row to input records
    ln = len(xllist)
           #the length of the list xllist
    if ln > 0: #リストに要素がない場合を排除 v1.01 2022/4/3
        ln_ = len(xllist[0])   #the number of the xllist's list[0]
        for i in range(0, ln):
            for j in range(0, ln_):
                sheet.cell(row=rn, column=j+fstcol).value = xllist[i][j]
            rn = rn + 1
            print('add a new transfer informatyon')
    else:
    	print(' xllist have no element!')
        
    #wb.save(wbN)


#fpychk_drecords#########################################################
"""
fpychk_drecords   :check doublue records
    重複データを別シートに抜き出す
v1.0
2022/3/30

@author: inoue
"""
def fpychk_drecords(wbN, sheetN):
    """
    check doublue records
    重複データを別シートに抜き出す
    Parameters
    ----------
    wbN : str
        Excelfile to check double data  '??_CowsHistory.xlsx'
    sheetN : str
        sheet name to check double data   '??Farm'

    Returns
    -------
    None.

    """
    #import chghistory
    #wbobj = chghistory.fpyopenxl(wbN, sheetN)
    #wb = wbobj[0]
    #sheet = wbobj[1]
    
    #excelfileのデータをlists'listにする
    xllists = fpyxllist_to_list(wbN,sheetN, 11)
    #value"0"のカラムflagをすべてのリストに追加する
    xllists_0 = fpyaddclm_to_lsts_lst(xllists, 0)
    #重複データのflagを0->1に変更する
    xllists_01 = fpyflag_dblrcd_1(xllists_0)
    #重複データのないlist
    xllists0 = fpydel_dblrcd(xllists_01, 11, 0)
    #重複していたデータのリスト
    xllists1 = fpydel_dblrcd(xllists_01, 11, 1)
   
    xllists0 = fpydelclm_frm_lsts_lst(xllists0, 11) 
    #col 'flag'の削除
    xllists1 = fpydelclm_frm_lsts_lst(xllists1, 11) 
    #col 'flag'の削除
    
    
    #シート名の変更
    fpychgSheetTitle(wbN, sheetN, sheetN + 'org')
    #振り分け用のシート　KTFarm　と　KTFarmout　を作成する。
    fpyNewSheet(wbN, sheetN, 'columns', 1)
    fpyNewSheet(wbN, sheetN + 'out', 'columns', 1)
    #データを振り分ける
    
    fpylisttoxls( xllists0, 1, wbN, sheetN)
    fpylisttoxls( xllists1, 1, wbN, sheetN + 'out')
    

#fpyreplace_str#########################################################
"""
fpyreplace_str : replace str to another str
    v1.0
    2022/7/12
    @author: jicc
    
"""
def fpyreplace_str(text, txt0, txt1):
    '''
    replace str to another str

    Parameters
    ----------
    text : str
     ex. abc\u3000def
    txt0 : str
        ex. \u3000
    txt1 : str
        ex. ' '

    Returns
    -------
    txt
    'abc def'

    '''

    txt = text.replace(txt0, txt1)
    
    return txt

#fpylstelemreplace_str#####################################################
"""
fpylstelemreplace_str : replace str to another str in a list's list
    v1.02
    2022/7/13
    @author: jicc
    
"""
def fpylstelemreplace_str(lst, elem, txt0, txt1):
    '''
    replace str to another str in a list's list

    Parameters
    ----------
    lst : list's list
     [[...], [...], ...]
     
    elem : int
        an element No of the target element to replace str 
    txt0 : str
        ex. \u3000
    txt1 : str
        ex. ' '

    Returns
    -------
    lst

    '''
    l = len(lst)
    for i in range(0, l):
        if lst[i][elem] == None:    #add if ~ else v1.01
            lst[i][elem] = '' #if lst[i][elem]...Noneとなる 
                              # lst[i][elem] -> '' に変更 v1.02
        else:
            lst[i][elem] = fpyreplace_str(lst[i][elem], txt0, txt1)

    return lst

#fpyselect_newrecords#######################################################
"""
fpyselect_newrecords   :select new records from transfer information
    異動情報から、新しいレコードを選択する
v1.0
2022/7/16

@author: inoue
"""
def fpyselect_newrecords(wbN, sheetN, ncol, idno):
    """
    select new records from transfer information
    異動情報から、新しいレコードを選択する
    Parameters
    ----------
    wbN : str
        Excelfile to check double data  'cowshistory.xlsx'
    sheetN : str
        sheet name of cowshistory   '??Farm'
    ncol :  int
        columns number of Excelfile's list
    idno : str
        ex. "0123456789"

    Returns
    -------
    lists' list　 [[title], [newrecords], [overlapped records]]

    """
    #import chghistory
    import nlbcs
    import time
    
    #excelfileのデータをlists'listにする
    xllists = fpyxllist_to_indlist(wbN, sheetN, ncol, idno)
    #氏名の全角空白'u\3000'を' 'に変換する
    xllists = fpylstelemreplace_str(xllists, 10, '\u3000', ' ')
   
    #個体識別情報検索画面のオープン
    driver = nlbcs.fpyopen_url("https://www.id.nlbc.go.jp/CattleSearch/search/agreement")
    nlbcs.fpyname_click(driver, "method:goSearch") 
    #個体識別番号 idno の情報を検索し、[[個体情報+異動情報], ...]
    #lists'list[[個体情報+異動情報], ...]を得る
    trs_inf = nlbcs.fpytrsinf_to_list(driver, idno)

    #lists' list [[title], [newrecords], [overlapped records]]
    trs_inf01 = [] #default
    
    #value"0"のカラムflagをすべてのリストに追加する
    trs_inf_0 = fpyaddclm_to_lsts_lst(trs_inf, 0)

    #xllistsにすでにあるlistのflagを0->1に変更する
    trs_inf_01 = fpyflag_dblrcd_1_(xllists, trs_inf_0)

    #list of new records
    trs_inf0 = fpydel_dblrcd(trs_inf_01, 10, 0)

    #list of overlapped records
    trs_inf1 = fpydel_dblrcd(trs_inf_01, 10, 1)

    #delete col 'flag'
    trs_inf0 = fpydelclm_frm_lsts_lst(trs_inf0, 10)

    #delete col 'flag'
    trs_inf1 = fpydelclm_frm_lsts_lst(trs_inf1, 10)
     
    trs_inf01.append(trs_inf0[0])  #[[title]]
    trs_inf01.append(trs_inf0[1:]) #[[title], [newrecords]]
    trs_inf01.append(trs_inf1)     #[[title], [newrecords], [overlapped records]] 
    print('trs_inf01')
    print(trs_inf01)
    
    time.sleep(3)
    nlbcs.fpydriver_quit(driver)
    
    return trs_inf01


"""
fpyselect_newrecords_s   :select new records from transfer information
    異動情報から、新しいレコードを選択する
    arguments 'wbN, sheetN' -> 'sheet' worksheetobject and 
    add arguments 'driver' Webdriver object
    
v1.0
2022/7/17

@author: inoue
"""
def fpyselect_newrecords_s(driver, sheet, ncol, idno):
    """
    select new records from transfer information
    異動情報から、新しいレコードを選択する
    Parameters
    ----------
    driver : webdriver.chrome.webdriver.WebDriver
        WebDriver object of selenium.webdriver.chrome.webdriver module
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    ncol :  int
        the number of columns of sheet(Excelfile's list)
    idno : str
        ex. "0123456789"

    Returns
    -------
    trs_inf01 : lists'list
    [[title], [[newrecord],..], [[overlapped record],..]] 

    """
    #import chghistory
    import nlbcs
    import time
    
    #excelfileのデータをlists'listにする
    xllists = fpyxllist_to_indlist_s(sheet, ncol, idno)

    xllists = fpylstelemreplace_str(xllists, 10, '\u3000', ' ')
        
    trs_inf = nlbcs.fpytrsinf_to_list(driver, idno)
    #time.sleep(3)
    #lists' list [[title], [[newrecord],..], [[overlapped record],..]] 
    trs_inf01 = [] #default
    #value"0"のカラムflagをすべてのリストに追加する
    trs_inf_0 = fpyaddclm_to_lsts_lst(trs_inf, 0)
    
    #xllistsにすでにあるlistのflagを0->1に変更する
    trs_inf_01 = fpyflag_dblrcd_1_(xllists, trs_inf_0)
    
    #list of new records
    trs_inf0 = fpydel_dblrcd(trs_inf_01, 10, 0)
    
    #list of overlapped records
    trs_inf1 = fpydel_dblrcd(trs_inf_01, 10, 1)
    
    #delete col 'flag'
    trs_inf0 = fpydelclm_frm_lsts_lst(trs_inf0, 10)
    
    #delete col 'flag'
    trs_inf1 = fpydelclm_frm_lsts_lst(trs_inf1, 10)
        
    trs_inf01.append(trs_inf0[0])  #[[title]]
    trs_inf01.append(trs_inf0[1:]) #[[title], [[newrecords],..]]
    trs_inf01.append(trs_inf1)     
    #[[title], [[newrecord],..], [[overlapped record],..]] 
    #print('trs_inf01')
    #print(trs_inf01)
    
    time.sleep(3)
    #nlbcs.fpydriver_quit(driver)
    
    return trs_inf01
    #trs_inf[1] : newrecords, trs_inf[2] : overlapped records
    
#fpynewtrs_inf_to_list#####################################################
"""
fpynewtrs_inf_to_list:
    compare original taransfer information with new information and
    separate new recors and overlapped records
    v1.0
    2022/7/22
    @author: jicc
    
"""
def fpynewtrs_inf_to_list(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1):
    """
    compare original taransfer information with new information and
    separate new recors and overlapped records

    Parameters
    ----------
    wbN0 : str
        Excelfile name of originaldata
        ex. "cowhistory.xlsx"
    sheetN0 : str
        sheet name
        ex. "MHFarm"
    colidno0 : int
        column number of 'idno0'(sheetN0 original data)
    
    wbN1 : str
        Excelfile name of new information
        ex. "??_cowslist.xlsx"
    sheetN1 : str
        sheet name
        ex. "cowslist"
    colidno1 : int
        column number of 'idno1' (sheetN1 new data)

    Returns
    -------
    trs_inf01 : lists'list
    [[newrecors'list], [overlappedrecords'list]]
    no title list

    """
    import nlbcs
    #import chghistory
    import time
    from selenium.common.exceptions import NoSuchElementException
    
    wb0 = fpyopenxl(wbN0, sheetN0)
    sheet0 = wb0[1]
    #max_row0 = sheet0.max_row
    
    wb1 = fpyopenxl(wbN1, sheetN1)
    sheet1 = wb1[1]
    max_row1 = sheet1.max_row
    
    trs_inf0 = [] #new records lists'list default
    trs_inf1 = [] #overlapped records lists'list default
    trs_inf01 = [] #all records which have searched

    driver = nlbcs.fpyopen_url("https://www.id.nlbc.go.jp/CattleSearch/search/agreement")
    nlbcs.fpyname_click(driver, "method:goSearch") 
    for row_num1 in range(2, max_row1 + 1):
        
        idno1 = fpygetCell_value(sheet1, row_num1, colidno1)
        
        try:
            tmp = fpyselect_newrecords_s(driver, sheet0, 11, idno1)
            #trs_inf0.append(tmp[0]) #columns list
            trs_inf0.append(tmp[1]) #new records list
            #trs_inf1.append(tmp[0]) #columns list
            trs_inf1.append(tmp[2]) #overlapped records list

        except NoSuchElementException:
             print("Error: " + idno1 + " not found")
                
    trs_inf01 = [trs_inf0, trs_inf1]
    #[[[newrecord],..], [[overlapped record],..]] 
    
    time.sleep(3)
    nlbcs.fpydriver_quit(driver)
    return trs_inf01

#fpynewtrs_inf_to_list_s#####################################################
"""
fpynewtrs_inf_to_list_s:
    compare original taransfer information with new information and
    separate new recors and overlapped records
    arguments 'wbN?, sheetN?' -> 'sheet?' worksheetobject version 
    v1.0
    2022/7/27
    @author: jicc
    
"""
def fpynewtrs_inf_to_list_s(sheet0, colidno0, sheet1, colidno1):
    """
    compare original taransfer information with new information and
    separate new recors and overlapped records

    Parameters
    ----------
    sheet0 : worksheet.worksheet.Worksheet
         worksheet object
    colidno0 : int
        column number of 'idno0'(sheet0 original data ex.MHFarm)
    sheet1 : worksheet.worksheet.Worksheet
         worksheet object
    colidno1 : int
        column number of 'idno1' (sheet1 new data ex. "cowslist")

    Returns
    -------
    trs_inf01 : lists'list
    [[newrecors'list], [overlappedrecords'list]]
    no title list

    """
    import nlbcs
    #import chghistory
    import time
    from selenium.common.exceptions import NoSuchElementException
    
    max_row1 = sheet1.max_row
    
    trs_inf0 = [] #new records lists'list default
    trs_inf1 = [] #overlapped records lists'list default
    trs_inf01 = [] #all records which have searched

    driver = nlbcs.fpyopen_url("https://www.id.nlbc.go.jp/CattleSearch/search/agreement")
    nlbcs.fpyname_click(driver, "method:goSearch") 
    for row_num1 in range(2, max_row1 + 1):
        
        idno1 = fpygetCell_value(sheet1, row_num1, colidno1)
        
        try:
            tmp = fpyselect_newrecords_s(driver, sheet0, 11, idno1)
            #trs_inf0.append(tmp[0]) #columns list
            trs_inf0.append(tmp[1]) #new records list
            #trs_inf1.append(tmp[0]) #columns list
            trs_inf1.append(tmp[2]) #overlapped records list

        except NoSuchElementException:
             print("Error: " + idno1 + " not found")
                
    trs_inf01 = [trs_inf0, trs_inf1]
    #[[[newrecord],..], [[overlapped record],..]] 
    
    time.sleep(3)
    nlbcs.fpydriver_quit(driver)
    return trs_inf01

#fpytrs_infs_to_xlsx########################################################
"""
fpytrs_infs_to_xlsx:
    search and save individual transfer informations to Excelfile
    v1.0
    2022/7/26
    @author: jicc
    
"""
def fpytrs_infs_to_xlsx(wbN0, sheetN0, wbN1, sheetN1, colidno1):
    """
    search and save individual transfer informations to Excelfile

    Parameters
    ----------
    wbN0 : str
        Excelfile name of originaldata
        ex. "cowhistory.xlsx"
    sheetN0 : str
        sheet name
        ex. "MHFarm"
    wbN1 : str
        Excelfile name of new idno information
        ex. "??_cowslist.xlsx"
    sheetN1 : str
        sheet name
        ex. "cowslist"
    colidno1 : int
        column number of 'idno1' (sheetN1 new data)

    Returns
    -------
    None.

    """
    import nlbcs
    #import chghistory
    import time
    from selenium.common.exceptions import NoSuchElementException
    
    wb0obj = fpyopenxl(wbN0, sheetN0) #[wb0, sheet0]
    wb0 = wb0obj[0] #ex. cowshistory.xlsx
    sheet0 = wb0obj[1] #ex. MHFarm
    #max_row0 = sheet0.max_row
    
    wb1obj = fpyopenxl(wbN1, sheetN1) #[wb1, sheet1]
    #wb1 = wb1obj[0] #ex. ??_cowslist.xlsx
    sheet1 = wb1obj[1] #ex. cowslist
    max_row1 = sheet1.max_row 
    
    driver = nlbcs.fpyopen_url("https://www.id.nlbc.go.jp/CattleSearch/search/agreement")
    nlbcs.fpyname_click(driver, "method:goSearch") 
    for row_num1 in range(2, max_row1 + 1):
        
        idno1 = fpygetCell_value(sheet1, row_num1, colidno1)
        
        try:

            nlbcs.fpytrsinf_to_xlsx(driver, idno1, sheet0)            
            
        except NoSuchElementException:
             print("Error: " + idno1 + " not found")
    
    wb0.save(wbN0)
    time.sleep(3)
    nlbcs.fpydriver_quit(driver)
    
#fpynewtrs_infs_to_xlsx######################################################    
"""
fpynewtrs_infs_to_xlsx : 
    search individual transfer informations  
    select new transfer informations
    input and save Excelfile
    v1.0
    2022/7/28 @author: jicc
    
"""

def fpynewtrs_infs_to_xlsx(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1):
    """
    search individual transfer informations  
    select new transfer informations
    input and save Excelfile    

    Parameters
    ----------
    wbN0 : str
        Excelfile name of originaldata
        ex. "cowhistory.xlsx"
    sheetN0 : str
        sheet name
        ex. "MHFarm"
    colidno0 : int
        column number of 'idno0' (sheetN0 )
    wbN1 : str
        Excelfile name of new idno information
        ex. "??_cowslist.xlsx"
    sheetN1 : str
        sheet name
        ex. "cowslist"
    colidno1 : int
        column number of 'idno1' (sheetN1 new data)

    Returns
    -------
    None.

    """

    #import chghistory
    
    wb0obj = fpyopenxl(wbN0, sheetN0)
    wb0 = wb0obj[0]
    sheet0 = wb0obj[1]
    #max_row0 = sheet0.max_row
    
    wb1obj = fpyopenxl(wbN1, sheetN1)
    sheet1 = wb1obj[1]
    #max_row1 = sheet1.max_row

    trs_inf01 = \
        fpynewtrs_inf_to_list_s(sheet0, colidno0, sheet1, colidno1)
        
    print('trs_inf01')
    print(trs_inf01)
        
    trs_inf0 = trs_inf01[0] #newrecords
    l0 = len(trs_inf0)
    if l0 > 0:
        for i in range(0, l0):
            
            fpylisttoxls_s(trs_inf0[i], 2, sheet0)
            
        wb0.save(wbN0)
           
######################################################################
"""
fpychghistoryReference:         reference of chbhistory's functions
ｖ1.1
2022/4/2
@author: jicc
"""
def fpychghistoryReference():
    
    print('-----chghistoryReference ---------------------------------------------------------v1.05------')
    print('**fpyopenxl(wbN, sheetN)')
    print('Excelfile wbN.xlsx　sheet sheetN Open ')
    print('.............................................................................................')
    print('**fpyopencsv_robj(csvN)')
    print('csvfile Open for Reader object')
    print('.............................................................................................')
    print('**fpyopencsv_rdata(csvN)')
    print('csvfile Open for Reader data')
    print('.............................................................................................')
    print('**fpyopencsv_w(csvN)')
    print('csvfile Open for Writer')
    print('.............................................................................................')
    print('**fpypdf_to_csv(filename, Path)')
    print('convert ****.pdf to ****.csv file')
    print('.............................................................................................')
    print('**fpySpdf_in_Dir_to_csv(Ext, Path)')
    print('ディレクトリ内の特定の拡張子(.pdf)を持つファイルを見つけcsvfile に変換する')
    print('.............................................................................................')
    print('**fpyCowHistory(csvorgN, csvoutN)')
    print('牛の個体情報.csvから、CowHistory.csv(changehistory\'s list )を作成する')
    print('.............................................................................................')
    print('**fpycsvlisttoxls(csvN, wbN, sheetN)')
    print('csvfileのデータをexcelfileに移行する')
    print('.............................................................................................')
    print('**fpycsvidNo_9to10( idNo )')
    print('idNo in a csvfile 9figures to 10figures')
    print('.............................................................................................')
    print('**fpydate_dottoslash( date )')
    print('date in a csvfile \'yyyy.mm.dd\' to \'yyyy/mm/dd\'')
    print('.............................................................................................')
    print('**fpyymd_csvtoCowsHistory_csv(Ext, Path, bckPath)')
    print('フォルダー内の個体履歴org(csv)をCowsHistory.csvに変更する')
    print('.............................................................................................')
    print('**fpyHistory_csvto_xlsx(Ext, Path, bckPath, wbN, sheetN)')
    print('フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する')
    print('.............................................................................................')
    print('**fpystrtodatetime( date )')
    print('str\'yyyy/mm/dd\'をdatetime に変換する')
    print('.............................................................................................')
    print('**fpyxlstrymdtodatetime(wbN, sheetN, col)')
    print('Excel cell \'yyyy/mm/dd\'をdatetimeに変換する')
    print('.............................................................................................')
    print('**fpyxllist_to_list(wbN, sheetN, ncol)')
    print('excelfileのリストを　lists\'　list にする')
    print('.............................................................................................')
    print('**fpyxllist_to_indlist(wbN, sheetN, ncol, idno)')
    print('get an individual lists\' list from excelfile\'s list')
    print('.............................................................................................')
    print('get an individual lists\' list from excelfile\'s list')
    print('**fpyxllist_to_indlist_s(sheet, ncol, idno)')
    print('arguments \'wbN, sheetN\' -> \'sheet\' worksheetobject version ')
    print('.............................................................................................')
    print('**fpyaddclm_to_lsts_lst(xllists, colv)')
    print('lists\'listに最終カラムを追加する')
    print('.............................................................................................')
    print('**fpydelclm_frm_lsts_lst(xllists, col)')
    print('lists\'listのカラムを削除する')
    print('.............................................................................................')
    print('**fpyflag_dblrcd_1_(xllists, trs_inf)')
    print('2つのlists\'list　listorgとlisttmpを比較し、')
    print('2つのlists\'listtmpの重複リストに　1（重複）でチェックを入れる')
    print('.............................................................................................')
    print('**fpydel_dblrcd(xllists, ,coln, colv)')
    print('lists\'listの重複リストの一つを削除する')
    print('.............................................................................................')
    print('**fpylisttoxls(xllist, fstcol, wbN, sheetN)')
    print('listのデータをexcelfileに移行する')
    print('....................................................................................')
    print('**fpychgSheetTitle(wbN, sheetN, sheetN1)')
    print('change ExcelSheet\'s title')
    print('....................................................................................')
    print('**fpyreplace_str(text, txt0, txt1)')
    print('replace str to another str')
    print('....................................................................................')
    print('**fpylstelemreplace_str(lst, elem, txt0, txt1)')
    print('replace str to another str in a list\'s list')
    print('....................................................................................')
    print('**fpyselect_newrecords(wbN, sheetN, ncol, idno)')
    print('select new records from transfer information')
    print('異動情報から、新しいレコードを選択する')
    print('....................................................................................')
    print('**fpyselect_newrecords_s(driver, sheet, ncol, idno)')
    print('select new records from transfer information')
    print('異動情報から、新しいレコードを選択する')
    print('arguments \'wbN, sheetN\' -> \'sheet\' worksheetobject and ')
    print('add arguments \'driver\' Webdriver object')
    print('....................................................................................')
    print('**fpynewtrs_inf_to_list(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1)')
    print('compare original taransfer information with new information and')
    print('separate new recors and overlapped records')
    print('....................................................................................')
    print('**fpytrs_infs_to_xlsx(wbN0, sheetN0, wbN1, sheetN1, colidno1)')
    print('search and save individual transfer informations to Excelfile')
    print('....................................................................................')
    print('**fpynewtrs_infs_to_xlsx(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1)')
    print('search individual transfer informations')
    print('select new transfer informations')
    print('input and save Excelfile ')
    print('--------------------------------------------------------------------2022/7/28　by jicc---------')
    
    
"""
fpyCowsHistoryManualfrmpdf:                        マニュアル
ｖ1.0
2022/1/11
@author: jicc
"""
def fpyCowsHistoryManualfrmpdf():
    
    print('-----CowsHistoryManual from pdffile--------------------------------------------v1.01-------')
    print(' ')
    print(' \"牛の個体情報検索サービス-個体識別番号の検索\"から個体の異動情報を検索し、')
    print('保存したpdffilesをcsvfileを介してExcelファイルにリスト化する。 ')
    print(' ')
    print('1.ディレクトリ内(..//CowsHistory)の特定の拡張子(.pdf)を持つファイルを見つけcsvfile に変換する')
    print('	MH_???_yyyymmdd.pdf -> ****.csv テーブル部分のデータ抽出')
    print('	MH_???_yyyymmdd.pdf -> ".\\pdforg\\"へ移行')
    print('   PS> python ps_fpyspdfindirtocsv_args.py Ext Path bckPath')
    print(' Ext: \.pdf, Path: .\\(カレントディレクトリ), bckPath: .\\pdforg')
    print(' ')
    print('2.フォルダー内の個体履歴org(csv)をCowsHistory.csvに変更する')
    print('変更後orgcsvfile を　別フォルダー(./csvorg)に移動する')
    print('MH_???_yyyymmdd.csv -> MH_???_yyyymmddH.csv')
    print('	MH_???_yyyymmdd.csv -> ".\\csvorg\\"へ移行')
    print('   PS> ps_fpyymd_csvtocowshistory_csv_args.py Ext Path bckPath')
    print(' Ext: \.csv, Path: .\\(カレントディレクトリ), bckPath: .\\csvorg')
    print(' ')
    print('3.フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する')
    print('移動後???H.csvを　別フォルダー(./csvhistory)に移動する')
    print('	MH_???_yyyymmddH.csv -> ".\\csvhistory\\"へ移行')
    print('   PS> ps_fpyhistory_csvto_xlsx_args.py Ext Path bckPath wbN sheetN')
    print(' Ext: \.csv, Path: .\\(カレントディレクトリ), bckPath: .\\csvhistory')
    print(' wbN: ..\\KT_CowsHistory.xlsx, sheetN:KTFarm')
    print(' ')
    print('4.??_CowsHistory.xlsx\/??Farm の　str\"yyyy\/mm\/dd\"を')
    print('datetimeに変換する')
    print('   PS> ps_fpyxlstrymdtodatetime_args.py wbN sheetN　col')
    print(' wbN: ..\\KT_CowsHistory.xlsx, sheetN:KTFarm, col: 3 and 9')
    print('---------------------------------------------------------------2022/3/6 by jicc---------')
    
"""
fpyCowsHistoryManualfrmweb                        マニュアル
ｖ1.0
2022/7/11
@author: jicc
"""
def fpyCowsHistoryManualfrmweb():
    
    print('-----CowsHistoryManual from web site------------------------------------------v1.0-------')
    print(' ')
    print('\"牛の個体情報検索サービス-個体識別番号の検索\"から個体の異動情報を検索し、 ')
    print('Excelファイルにリスト化する。 ')
    print('web -> csvfile ->Excelfile ')
    print(' ')
    print('1. ABFarmの個体リスト(AB_cowslist.xlsx/ABFarm)から、個体識別番号(colum2 idno)によって、')
    print('個体情報+異動情報を検索し、リストにし、idno_ymd.csv fileに保存する')
    print('   PS> ps_fpyindtrsinf_to_csv_args.py wbN sheetN')
    print(' wbN : AB_cowslist.xlsx, sheetN : cowslist')
    print('牛の個体情報検索サービス-個体識別番号の検索')
    print(' url : https://www.id.nlbc.go.jp/CattleSearch/search/agreement')
    print(' ')
    print('2.フォルダー内のidno_ymd.csvをcowshistory.xlsx/ABFarmに移動する')
    print('移動後idno_ymd.csvを　別フォルダー(./csvhistory)に移動する')
    print('	idno_ymd.csv -> ".\\csvhistory\\"へ移動')
    print('   PS> python ps_fpyhistory_csvto_xlsx_args.py Ext Path bckPath wbN sheetN')
    print(' Ext: \.csv, Path: .\\(カレントディレクトリ), bckPath: .\\csvhistory')
    print(' wbN: (..\\)cowshistory.xlsx, sheetN:ABFarm')
    print(' ')
    print('3.cowshistory.xlsx\/ABFarm の　str\"yyyy\/mm\/dd\"を')
    print('datetimeに変換する')
    print('   PS> ps_fpyxlstrymdtodatetime_args.py wbN sheetN　col')
    print(' wbN: ..\\KT_CowsHistory.xlsx, sheetN:KTFarm, col: 3 and 9')
    print('---------------------------------------------------------------2022/7/11 by jicc---------')
    
    
"""
fpyCowsHistoryTools:                        tools
ｖ1.0
2022/7/29
@author: jicc
"""
def fpyCowsHistoryTools():
    
    print('-----CowsHistoryTools---------------------------------------------------------v2.00-------')
    print('牛の個体情報検索サービス 個体識別番号の検索から個体の異動情報を検索し、')
    print('Excelファイルにリスト化するための　Tool集')
    print(' ')
    print('#fpytrs_infs_to_xlsx(wbN0, sheetN0, wbN1, sheetN1, colidno1)')
    print('個体リスト AB_cowslist/cowslistのidnoから個体異動情報を検索する')
    print('個体情報リスト cowshistory.xlsx/ABFarmに新規または追加入力する')
    print('   PS> ps_fpytrs_infs_to_xlsx_args.py wbN0 sheetN0 wbN1 sheetN1 colidno1')
    print(' wbN0 : cowshistory.xlsx, sheetN0 : ABFarm, ')
    print(' wbN1 : AB_cowslist.xlsx, sheetN1 : cowslist, colidno1 : 2 (column number fo idno1)')
    print(' ')
    print('#fpynewtrs_infs_to_xlsx(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1)')
    print('個体リスト AB_cowslist/のidnoから個体異動情報を検索する')
    print('個体情報リスト cowshistory.xlsx/ABFarmにない新しい情報を抽出する')
    print(' cowshistory.xlsx/ABFarmに追加入力する')
    print('   PS> ps_fpynewtrs_infs_to_xlsx_args.py wbN0 sheetN0 colidno0 wbN1 sheetN1 colidno1')
    print('  wbN0 : cowshistory.xlsx, sheetN0 : ABFarm, colidno0 : 2 (column number fo idno0), ')
    print(' wbN1 : AB_cowslist.xlsx, sheetN1 : cowslist, colidno1 : 2 (column number fo idno1)')
    print(' ')
    print('#fpychk_drecords(wbN, sheetN)')
    print('Excel個体情報リスト cowshistory/ABFarmの重複データをを削除する')
    print('   PS> ps_fpychk_drecords_args.py wbN sheetN')
    print(' wbN: ..\\cowshistory.xlsx, sheetN:ABFarm')
    print('---------------------------------------------------------------2022/7/29by jicc---------')    
